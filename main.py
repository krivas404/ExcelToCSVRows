import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math
import csv
import settings as s


file_name = s.file_name
new_file = s.new_file
sheet_name = s.sheet_name

wb = load_workbook(filename=file_name, read_only=True)
sheet = wb[sheet_name]
value_to_roundnum = s.value_to_roundnum

def GetValuePlusOne(value):
    value = int(value)
    value += 1
    if value == 10:
        return "1"
    return str(value)

# def RoundUp(num, part_num, point, value):
#     i = int(num[point+value]) + 1
#     if i == 10:


def SaveNumWithoutDistortion(num):
    num = num * 1000
    result = int(num // 10)
    filler = ''
    if num < 100:
        filler = '0.0'
    if num % 10 >= 5:
        result = int(num // 10 + 1)
    if result < 0.1:
        


    string_result = str(result)
    saved_num = string_result[:-2] +'.' + string_result[-2:]
    return saved_num


# def RoundNum(num, value):
#     num = str(num)
#     point = num.index('.')
#     if len(num[point:]) <= 3:
#         return num
#     start_of_num = num[:point + value]
#     end_of_num = num[point + value + 1:]
#     number_to_check = end_of_num[0]
#     num_to_add = num[point+value]
#     if end_of_num and int(number_to_check) > 4:
#         num_to_add = GetValuePlusOne(num_to_add)
#     start_of_num = start_of_num + num_to_add
#     return start_of_num


def GetRowToWrite(row):
    row1_to_write = row[1].value
    if row[5].value == '100':
        row2_to_write = RoundNum(row[6].value / 100, 2)
        return (row1_to_write, row2_to_write)
    row2_to_write = row[6].value
    if type(row2_to_write) == float:
        row2_to_write = RoundNum(row2_to_write, 2)
    return (row1_to_write, row2_to_write)


def main():
    print('started')
    val1 = sheet['B1'].value
    print('val1 =', val1)
    i = 0
    with open(new_file, 'w', newline='') as csv_file:
        csvwriter = csv.writer(csv_file, dialect='excel', delimiter=';')
        for row in sheet.rows:
            if i == 0:
                i += 1
                continue
            row_to_write = GetRowToWrite(row)
            csvwriter.writerow(row_to_write)
            i += 1


if __name__ == "__main__":
    main()
